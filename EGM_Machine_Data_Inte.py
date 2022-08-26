import openpyxl as xl
import pandas as pd

# EGM_teplate_loc = r'C:\Users\ALiu\Desktop\Andy\16. Projects\Application\EGM-TEST\EGM\test all'
# EGM_template = 'Data Templates - EGM - WK '
# week = '34'
# weekly_template = EGM_template + week + ".xlsx"
# weekly_template_EGM = os.path.join(EGM_teplate_loc, weekly_template)


EGM_TEMPLATE=r'C:\Users\ALiu\Desktop\Andy\16. Projects\Application\EGM-TEST\EGM\test all\Data Templates - EGM - WK 34.xlsx'
test=r'C:\Users\ALiu\Desktop\Andy\16. Projects\Application\EGM-TEST\EGM\test all\44444.xlsx'
machine_data=r'C:\Users\ALiu\Desktop\Andy\16. Projects\Application\EGM-TEST\EGM\test all\Machine Data.xlsx'
new_machinedata = pd.DataFrame()


# Source_machine_sheet = pd.read_excel(EGM_TEMPLATE,sheet_name='Machine Data - Source')
# floor_analysis_sheet=pd.read_excel(EGM_TEMPLATE,sheet_name='Floor Analysis - Data')
# # machine_data_sheet=pd.read_excel(machine_data,sheet_name='Sheet1')
#
#
# df_source = pd.DataFrame(Source_machine_sheet)
# df_floor_analysis=pd.DataFrame(floor_analysis_sheet)
# # df_machine_data=pd.DataFrame(machine_data_sheet)
#
# row=len(df_source.index)
# df_floor_analysis=df_floor_analysis.iloc[:row]
# # df_machine_data=df_machine_data.append(df_floor_analysis)
# # df_machine_data.to_excel(machine_data,index=False)
#
# #temp file
#
# row=len(df_source.index)
# df_floor_analysis=df_floor_analysis.iloc[:row]
# df_floor_analysis.to_excel(test,header=False,index=False)

#copy temp file to machine data

wb1 = xl.load_workbook(filename=test)
ws1 = wb1.worksheets[0]
# print(source_filename)
# print(wb1.sheetnames)

# # opening the destination excel file

target_file_name_loc = machine_data

wb2 = xl.load_workbook(filename=target_file_name_loc)
ws2 = wb2['Sheet1']


# calculate total number of rows and
# columns in source excel file
mr1 = ws1.max_row
mc1 = ws1.max_column

mr2 = ws2.max_row
mc2 = ws2.max_column

# # copying the cell values from source
# # excel file to destination excel file
for i in range(1, mr1 + 1):
    for j in range(1, mc1 + 1):
        # reading cell value from source excel file
        c = ws1.cell(row=i, column=j)
        # writing the read value to destination excel file
        ws2.cell(row=mr2+i, column=j).value = c.value
# saving the destination excel file
wb2.save(target_file_name_loc)
print("well done!!!")
