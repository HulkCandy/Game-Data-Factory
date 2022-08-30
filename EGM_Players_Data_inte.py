import openpyxl as xl
import pandas as pd


def EGM_inte_Playerdata():
    EGM_TEMPLATE=r'C:\Users\ALiu\Desktop\Andy\16. Projects\Application\EGM-TEST\EGM\test all\Data Templates - EGM - WK 34.xlsx'
    player_temp=r'C:\Users\ALiu\Desktop\Andy\16. Projects\Application\EGM-TEST\EGM\test all\Player_Temp.xlsx'
    machine_data=r'C:\Users\ALiu\Desktop\Andy\16. Projects\Application\EGM-TEST\EGM\test all\EGM - Loyalty Data.txt'
    all_data = r'C:\Users\ALiu\Desktop\Andy\16. Projects\Application\EGM-TEST\EGM\test all\EGM - Loyalty Data.txt'
    player_temp_txt = r'C:\Users\ALiu\Desktop\Andy\16. Projects\Application\EGM-TEST\EGM\test all\Player_temp.txt'


    Source_player_sheet = pd.read_excel(EGM_TEMPLATE,sheet_name='Player Data - Source')
    player_sheet=pd.read_excel(EGM_TEMPLATE,sheet_name='Loyalty Data - Report')
    Uncarded_sheet = pd.read_excel(EGM_TEMPLATE, sheet_name='Uncarded Data - Report')

    df_source = pd.DataFrame(Source_player_sheet)
    df_player=pd.DataFrame(player_sheet)
    df_Uncarded = pd.DataFrame(Uncarded_sheet)

    row=len(df_source.index)
    df_player=df_player.iloc[:row]
    print(df_player)
    df_player=df_player.append(df_Uncarded)
    df_player.iloc[:,4] = pd.to_datetime(df_player.iloc[:,4])
    df_player.iloc[:,4]=df_player.iloc[:,4].dt.strftime('%d/%m/%Y')

    df_player.to_csv(player_temp_txt, header=None, index=None, sep='\t', mode='a')

    ######combining data

    inputFile = player_temp_txt
    # Opening the given file in read-only mode.
    readFile = open(inputFile, "r")

    # output text file path
    outputFile = all_data
    # Opening the output file in write mode.
    outFile = open(outputFile, "r")

    # Read the above read file lines using readlines()
    ReadFileLines = readFile.readlines()
    writeFile = open(outputFile, "a")
    writeFile.write("\n")

    # Traverse in each line of the read text file
    for i in range(0, len(ReadFileLines)):
        writeFile.write(ReadFileLines[i])

    # Closing the write file
    writeFile.close()

    # Closing the read file
    readFile.close()






    # #copy temp file to machine data
    #
    # wb1 = xl.load_workbook(filename=test)
    # ws1 = wb1.worksheets[0]
    # # print(source_filename)
    # # print(wb1.sheetnames)
    #
    # # # opening the destination excel file
    #
    # target_file_name_loc = machine_data
    #
    # wb2 = xl.load_workbook(filename=target_file_name_loc)
    # ws2 = wb2['Sheet1']
    #
    #
    # # calculate total number of rows and
    # # columns in source excel file
    # mr1 = ws1.max_row
    # mc1 = ws1.max_column
    #
    # mr2 = ws2.max_row
    # mc2 = ws2.max_column
    #
    # # # copying the cell values from source
    # # # excel file to destination excel file
    # for i in range(1, mr1 + 1):
    #     for j in range(1, mc1 + 1):
    #         # reading cell value from source excel file
    #         c = ws1.cell(row=i, column=j)
    #         # writing the read value to destination excel file
    #         ws2.cell(row=mr2+i, column=j).value = c.value
    # # saving the destination excel file
    # wb2.save(target_file_name_loc)

if __name__ == '__main__':

    EGM_inte_Playerdata()