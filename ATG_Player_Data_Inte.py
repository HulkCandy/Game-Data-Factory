import openpyxl as xl
import pandas as pd

def ATG_inte_playerdata():
    ATG_TEMPLATE=r'C:\Users\ALiu\Desktop\Andy\16. Projects\Application\EGM-TEST\ATG\test all\Data Templates - ATG - WK 35.xlsx'
    test=r'C:\Users\ALiu\Desktop\Andy\16. Projects\Application\EGM-TEST\ATG\test all\Player_temp.xlsx'
    loyal_data=r'C:\Users\ALiu\Desktop\Andy\16. Projects\Application\EGM-TEST\ATG\test all\ATG Source Data - Loyalty.xlsx'



    Source_player_sheet = pd.read_excel(ATG_TEMPLATE,sheet_name='Player Data - Source')
    player_transformed_sheet=pd.read_excel(ATG_TEMPLATE,sheet_name='Player data transformed')
    Uncarded = pd.read_excel(ATG_TEMPLATE, sheet_name='Uncarded Data - Report')
    # loyal_data_sheet=pd.read_excel(loyal_data,sheet_name='Gaming Data')
    # Player_temp=pd.read_excel(test,sheet_name='Sheet1')


    Source_player_sheet = pd.DataFrame(Source_player_sheet)
    player_transformed_sheet=pd.DataFrame(player_transformed_sheet)
    uncarded_sheet = pd.DataFrame(Uncarded)
    # loyal_data_sheet=pd.DataFrame(loyal_data_sheet)
    # player_temp_sheet = pd.DataFrame(Player_temp)

    row=len(Source_player_sheet.index)
    player_transformed_sheet=player_transformed_sheet.iloc[:row]
    # uncarded_sheet=uncarded_sheet.iloc[0:8]
    print(uncarded_sheet)
    df=pd.concat([player_transformed_sheet,uncarded_sheet],axis=0,ignore_index=True)
    # df=player_transformed_sheet.append(uncarded_sheet)
    # df_machine_data = pd.concat(player_transformed_sheet)
    # print(df_machine_data)
    # df.to_excel(test,index=False,header=False)
    df.to_excel(test)



    #temp file

    # row=len(df_source.index)
    # df_floor_analysis=df_floor_analysis.iloc[:row]
    # df_floor_analysis.to_excel(test,header=False,index=False)
    #
    # #copy temp file to machine data
    #
    # wb1 = xl.load_workbook(filename=test)
    # ws1 = wb1.worksheets[0]
    # # print(source_filename)
    # # print(wb1.sheetnames)
    #
    # # # opening the destination excel file
    #
    # target_file_name_loc = machine_data
    #
    # wb2 = xl.load_workbook(filename=target_file_name_loc)
    # ws2 = wb2['Sheet1']
    #
    #
    # # calculate total number of rows and
    # # columns in source excel file
    # mr1 = ws1.max_row
    # mc1 = ws1.max_column
    #
    # mr2 = ws2.max_row
    # mc2 = ws2.max_column
    #
    # # # copying the cell values from source
    # # # excel file to destination excel file
    # for i in range(1, mr1 + 1):
    #     for j in range(1, mc1 + 1):
    #         # reading cell value from source excel file
    #         c = ws1.cell(row=i, column=j)
    #         # writing the read value to destination excel file
    #         ws2.cell(row=mr2+i, column=j).value = c.value
    # # saving the destination excel file
    # wb2.save(target_file_name_loc)

if __name__ == '__main__':
    ATG_inte_playerdata()