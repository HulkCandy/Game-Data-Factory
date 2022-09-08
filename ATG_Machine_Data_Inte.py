import openpyxl as xl
import pandas as pd



def ATG_inte_machinedata():
    ATG_TEMPLATE=r'C:\Users\ALiu\Desktop\Andy\16. Projects\Application\EGM-TEST\ATG\test all\Data Templates - ATG - WK 35.xlsx'
    test=r'C:\Users\ALiu\Desktop\Andy\16. Projects\Application\EGM-TEST\ATG\test all\Machine_temp.xlsx'
    machine_data=r'C:\Users\ALiu\Desktop\Andy\16. Projects\Application\EGM-TEST\ATG\test all\ATG Source data - Game.xlsx'
    new_machinedata = pd.DataFrame()


    Source_machine_sheet = pd.read_excel(ATG_TEMPLATE,sheet_name='Combined report')

    df_source = pd.DataFrame(Source_machine_sheet)

    df_floor_analysis=df_source.iloc[2:17]
    # print(df_floor_analysis)
    df_floor_analysis.iloc[:,0] = pd.to_datetime(df_floor_analysis.iloc[:,0])
    df_floor_analysis.iloc[:,0]=df_floor_analysis.iloc[:,0].dt.strftime('%d/%m/%Y')

    df_floor_analysis.to_excel(test,index=False,header=False)

    #temp file


    # copy temp file to machine data

    wb1 = xl.load_workbook(filename=test)
    ws1 = wb1.worksheets[0]
    # print(source_filename)
    # print(wb1.sheetnames)

    # # opening the destination excel file

    target_file_name_loc = machine_data

    wb2 = xl.load_workbook(filename=target_file_name_loc)
    ws2 = wb2['Source Data']


    # calculate total number of rows and
    # columns in source excel file
    mr1 = ws1.max_row
    mc1 = ws1.max_column

    mr2 = ws2.max_row
    mc2 = ws2.max_column

    # # copying the cell values from source
    # # excel file to destination excel file
    for i in range(1, mr1 + 1):
        for j in range(1, mc1 + 1):
            # reading cell value from source excel file
            c = ws1.cell(row=i, column=j)
            # writing the read value to destination excel file
            ws2.cell(row=mr2+i, column=j).value = c.value
    # saving the destination excel file
    wb2.save(target_file_name_loc)


if __name__ == '__main__':
    ATG_inte_machinedata()