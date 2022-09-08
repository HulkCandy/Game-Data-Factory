import os
from datetime import timedelta, datetime
import pandas as pd
import datetime as dt
import shutil
import regex as re
import openpyxl as xl
import xlsxwriter


# def validat(dates):
#     dates=dates
#     print("function is working")
#     print(dates)


# convert date to actual date
def ATG_account(dates):
    account_original = r'C:\Users\ALiu\Desktop\Andy\16. Projects\Application\EGM-TEST\ATG\Account'
    account_target = r'C:\Users\ALiu\Desktop\Andy\16. Projects\Application\EGM-TEST\ATG\Account - Test'
    dates=dates
    EGM_teplate_loc = r'C:\Users\ALiu\Desktop\Andy\16. Projects\Application\EGM-TEST\ATG\test all'
    EGM_template = 'Data Templates - ATG - WK '
    week = '35'
    weekly_template=EGM_template+week+".xlsx"
    weekly_template_EGM=os.path.join(EGM_teplate_loc, weekly_template)
    original=account_original
    target=account_target
    for i in range(len(dates)):
        dates[i] = datetime.strptime(dates[i], '%Y-%m-%d')
        dates[i] = pd.to_datetime(dates[i]) + pd.DateOffset(days=1)

    date_list = [t.strftime("%Y-%m-%d") for t in dates]

    dir_list = os.listdir(original)

    # print(dir_list)

    for i in range(len(date_list)):
        date_list[i] = date_list[i].replace('-', '')

    date_list = ["AccountingATG_{0}.txt".format(date) for date in date_list]

    # print(date_list)


    for i in range(len(date_list)):
        original = os.path.join(original, date_list[i])
        target = os.path.join(target, date_list[i])
        if os.path.isfile(original):
            shutil.copy(original, target)
            original = os.path.split(original)
            original = original[0]
            target = os.path.split(target)
            target = target[0]
    # print(target)


    target_list = os.listdir(target)
    # print(target_list)


    # clean machine data start
    colspecs = (
        (0, 16),
        (16, 29),
        (29, 46),
        (46, 87),
        (87, 120),
        (120, 133),
        (133, 184),
        (184, 206),
        (206, 233),
        (233, 274),
        (274, 287),
        (287, 296)
    )

    data_egm_list = []

    for filename in os.listdir(target):
        with open(os.path.join(target, filename), 'r', encoding='utf16') as fin:
            data = fin.read().splitlines(True)
        with open(os.path.join(target, filename), 'w', encoding='utf16') as fout:
            fout.writelines(data[2:])
            # fout.writelines(data[:-2])
            # fout.close()

    for filename in os.listdir(target):
        df = pd.read_fwf(os.path.join(target, filename), colspecs=colspecs, encoding='utf16')
        df = df.iloc[1:]
        df = df.iloc[:-2]
        # df.to_excel(r'C:\Users\ALiu\Desktop\Andy\16. Projects\Application\EGM-TEST\EGM\test all\test.xlsx',index=False)
        data_egm_list.append(df)

    # print(data_egm_list)
    final_egm_data = pd.concat(data_egm_list)
    # print(final_egm_data.dtypes)
    final_egm_data['Accounting Date'] = pd.to_datetime(final_egm_data['Accounting Date']).dt.date
    for col in final_egm_data.columns[8:]:
        final_egm_data[col] = pd.to_numeric(final_egm_data[col], errors='coerce')

    account_source_file_name = 'Account test.xlsx'
    account_source_file_name_loc = r'C:\Users\ALiu\Desktop\Andy\16. Projects\Application\EGM-TEST\ATG\test all'

    account_source_file_name_full = os.path.join(account_source_file_name_loc, account_source_file_name)
    final_egm_data.to_excel(account_source_file_name_full, header=False, index=False)
    # data clean and concat end

    # copy cleaned data into template

    # importing openpyxl module


    # opening the source excel file
    source_filename = account_source_file_name_full
    wb1 = xl.load_workbook(filename=source_filename)
    ws1 = wb1.worksheets[0]
    # print(source_filename)
    # print(wb1.sheetnames)


    # # opening the destination excel file
    account_target_file_name = EGM_template + week + '.xlsx'
    account_target_file_name_loc = os.path.join(EGM_teplate_loc, account_target_file_name)
    # print(account_target_file_name_loc)
    wb2 = xl.load_workbook(filename=account_target_file_name_loc)
    ws2=wb2['Machine Data - Source']
    # target_list1 = os.listdir(EGM_teplate_loc)
    # print(target_list1)

    # calculate total number of rows and
    # columns in source excel file
    mr = ws1.max_row
    mc = ws1.max_column

    # # copying the cell values from source
    # # excel file to destination excel file
    for i in range(1, mr + 1):
        for j in range(1, mc + 1):
            # reading cell value from source excel file
            c = ws1.cell(row=i, column=j)

            # writing the read value to destination excel file
            ws2.cell(row=i+1, column=j).value = c.value
    # saving the destination excel file
    wb2.save(account_target_file_name_loc)



def ATG_Players(dates):
    player_original = r'C:\Users\ALiu\Desktop\Andy\16. Projects\Application\EGM-TEST\ATG\Player'
    player_target = r'C:\Users\ALiu\Desktop\Andy\16. Projects\Application\EGM-TEST\ATG\Player - Test'
    dates = dates
    original=player_original
    target=player_target
    EGM_teplate_loc = r'C:\Users\ALiu\Desktop\Andy\16. Projects\Application\EGM-TEST\ATG\test all'
    EGM_template = 'Data Templates - ATG - WK '
    week = '35'
    weekly_template=EGM_template+week+".xlsx"
    weekly_template_EGM=os.path.join(EGM_teplate_loc, weekly_template)

    for i in range(len(dates)):
        dates[i] = datetime.strptime(dates[i], '%Y-%m-%d')
        dates[i] = pd.to_datetime(dates[i]) + pd.DateOffset(days=1)

    date_list = [t.strftime("%Y-%m-%d") for t in dates]


    dir_list = os.listdir(original)

    # print(dir_list)

    for i in range(len(date_list)):
        date_list[i] = date_list[i].replace('-', '')

    date_list = ["PlayerATG_{0}.txt".format(date) for date in date_list]

    # print(date_list)

# copy files from original to target

    for i in range(len(date_list)):
        original = os.path.join(original, date_list[i])
        target = os.path.join(target, date_list[i])
        if os.path.isfile(original):
            shutil.copy(original, target)
            original = os.path.split(original)
            original = original[0]
            target = os.path.split(target)
            target = target[0]
        else:
            print("shit")
    # print(target)


    target_list = os.listdir(target)
    # print(target_list)


    # clean machine data start
    colspecs = (
        (0, 10),
        (10, 47),
        (47, 62),
        (62, 113),
        (113, 130),
        (130, 142),
        (142, 164),
        (164, 186),
        (186, 208),
        (208, 222),
        (222, 242),
        (242, 254),
        (254, 285),
        (285, 316),
        (316, 323)

    )

    data_egm_list = []

    for filename in os.listdir(target):
        with open(os.path.join(target, filename), 'r', encoding='utf16') as fin:
            data = fin.read().splitlines(True)
        with open(os.path.join(target, filename), 'w', encoding='utf16') as fout:
            fout.writelines(data[2:])
            # fout.writelines(data[:-2])
            # fout.close()

    for filename in os.listdir(target):
        df = pd.read_fwf(os.path.join(target, filename), colspecs=colspecs, encoding='utf16')
        df = df.iloc[1:]
        df = df.iloc[:-2]
        # df.to_excel(r'C:\Users\ALiu\Desktop\Andy\16. Projects\Application\EGM-TEST\EGM\test all\test.xlsx',index=False)
        data_egm_list.append(df)

    # print(data_egm_list)
    final_egm_data = pd.concat(data_egm_list)
    final_egm_data['Accounting']=pd.to_datetime(final_egm_data['Accounting'], format='%d/%m/%Y').dt.date
    for col in final_egm_data.columns[6:12]:
        final_egm_data[col] = pd.to_numeric(final_egm_data[col], errors='coerce')

    # account_source_file_name = 'test.xlsx'
    account_source_file_name_loc = r'C:\Users\ALiu\Desktop\Andy\16. Projects\Application\EGM-TEST\ATG\test all'
    player_source_file_name = 'Player test.xlsx'

    account_source_file_name_full = os.path.join(account_source_file_name_loc, player_source_file_name)
    final_egm_data.to_excel(account_source_file_name_full, header=False, index=False)
    # data clean and concat end

    # copy cleaned data into template

    # importing openpyxl module


    # opening the source excel file
    source_filename = account_source_file_name_full
    wb1 = xl.load_workbook(filename=source_filename)
    ws1 = wb1.worksheets[0]
    # print(source_filename)
    # print(wb1.sheetnames)


    # # opening the destination excel file
    account_target_file_name = EGM_template + week + '.xlsx'
    account_target_file_name_loc = os.path.join(EGM_teplate_loc, account_target_file_name)
    # print(account_target_file_name_loc)
    wb2 = xl.load_workbook(filename=account_target_file_name_loc)
    ws2=wb2['Player Data - Source']
    # target_list1 = os.listdir(EGM_teplate_loc)
    # print(target_list1)

    # calculate total number of rows and
    # columns in source excel file
    mr = ws1.max_row
    mc = ws1.max_column

    # # copying the cell values from source
    # # excel file to destination excel file
    for i in range(1, mr + 1):
        for j in range(1, mc + 1):
            # reading cell value from source excel file
            c = ws1.cell(row=i, column=j)

            # writing the read value to destination excel file
            ws2.cell(row=i+1, column=j).value = c.value
    # saving the destination excel file
    wb2.save(account_target_file_name_loc)

# def EGM_integrat(weekly_template_EGM):
#     data = pd.read_excel(weekly_template_EGM,sheet_name='Uncarded Data - Report',header=None)
#     df = pd.DataFrame(data)
#     print(df)



    # wb2 = xl.load_workbook(filename=weekly_template_EGM)
    # ws2=wb2['Floor Analysis - Data']
    # mr = ws2.max_row
    # mc = ws2.max_column
    #
    # # # copying the cell values from source
    # # # excel file to destination excel file
    # for i in range(1, mr + 1):
    #     for j in range(1, mc + 1):
    #         # reading cell value from source excel file
    #         c = ws1.cell(row=i, column=j)
    #
    #         # writing the read value to destination excel file
    #         ws2.cell(row=i+1, column=j).value = c.value





if __name__ == '__main__':

    dates = ['2022-08-29', '2022-08-30', '2022-08-31', '2022-09-01', '2022-09-02', '2022-09-03', '2022-09-04']
    # account_original = r'C:\Users\ALiu\Desktop\Andy\16. Projects\Application\EGM-TEST\EGM\Account'
    # account_target = r'C:\Users\ALiu\Desktop\Andy\16. Projects\Application\EGM-TEST\EGM\Account-test'
    # player_original = r'C:\Users\ALiu\Desktop\Andy\16. Projects\Application\EGM-TEST\EGM\Players'
    # player_target = r'C:\Users\ALiu\Desktop\Andy\16. Projects\Application\EGM-TEST\EGM\Players - test'
    # date_list = []
    # EGM_teplate_loc = r'C:\Users\ALiu\Desktop\Andy\16. Projects\Application\EGM-TEST\EGM\test all'
    # EGM_template = 'Data Templates - EGM - WK '
    # week = '34'
    # weekly_template=EGM_template+week+".xlsx"
    # weekly_template_EGM=os.path.join(EGM_teplate_loc, weekly_template)

    # account_source_file_name = 'test.xlsx'
    # account_source_file_name_loc = r'C:\Users\ALiu\Desktop\Andy\16. Projects\Application\EGM-TEST\EGM\test all'
    # player_source_file_name = 'player test.xlsx'

    # ATG_account(dates)
    # ATG_Players(dates)
    # EGM_integrat(weekly_template_EGM)



