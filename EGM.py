import os
from datetime import timedelta, datetime
import pandas as pd
import datetime as dt
import shutil
import regex as re
import openpyxl as xl
import xlsxwriter

dates = ['2022-08-08', '2022-08-09', '2022-08-10', '2022-08-11', '2022-08-12', '2022-08-13', '2022-08-14']
date_list = []
EGM_teplate_loc = r'C:\Users\ALiu\Desktop\Andy\16. Projects\Application\EGM-TEST\EGM\test all'
EGM_template = 'Data Templates - EGM - WK '
week = '34'
account_source_file_name = 'test.xlsx'
account_source_file_name_loc = r'C:\Users\ALiu\Desktop\Andy\16. Projects\Application\EGM-TEST\EGM\test all'

original = r'C:\Users\ALiu\Desktop\Andy\16. Projects\Application\EGM-TEST\EGM\Account'
target = r'C:\Users\ALiu\Desktop\Andy\16. Projects\Application\EGM-TEST\EGM\Account-test'

# convert date to actual date

for i in range(len(dates)):
    dates[i] = datetime.strptime(dates[i], '%Y-%m-%d')
    dates[i] = pd.to_datetime(dates[i]) + pd.DateOffset(days=1)

date_list = [t.strftime("%Y-%m-%d") for t in dates]

dir_list = os.listdir(original)

# print(dir_list)

for i in range(len(date_list)):
    date_list[i] = date_list[i].replace('-', '')

date_list = ["Accouting_{0}.txt".format(date) for date in date_list]

# print(date_list)


for i in range(len(date_list)):
    original = os.path.join(original, date_list[i])
    target = os.path.join(target, date_list[i])
    if os.path.isfile(original):
        shutil.copy(original, target)
        original = os.path.split(original)
        original = original[0]
        target = os.path.split(target)
        target = target[0]
# print(target)


target_list = os.listdir(target)
# print(target_list)


# clean machine data start
colspecs = (
    (0, 16),
    (16, 29),
    (29, 46),
    (46, 87),
    (87, 120),
    (120, 133),
    (133, 184),
    (184, 206),
    (206, 233),
    (233, 274),
    (274, 287),
    (287, 296)
)

data_egm_list = []

for filename in os.listdir(target):
    with open(os.path.join(target, filename), 'r', encoding='utf16') as fin:
        data = fin.read().splitlines(True)
    with open(os.path.join(target, filename), 'w', encoding='utf16') as fout:
        fout.writelines(data[2:])
        # fout.writelines(data[:-2])
        # fout.close()

for filename in os.listdir(target):
    df = pd.read_fwf(os.path.join(target, filename), colspecs=colspecs, encoding='utf16')
    df = df.iloc[1:]
    df = df.iloc[:-2]
    # df.to_excel(r'C:\Users\ALiu\Desktop\Andy\16. Projects\Application\EGM-TEST\EGM\test all\test.xlsx',index=False)
    data_egm_list.append(df)

# print(data_egm_list)
final_egm_data = pd.concat(data_egm_list)
# print(final_egm_data)
account_source_file_name_full = os.path.join(account_source_file_name_loc, account_source_file_name)
final_egm_data.to_excel(account_source_file_name_full, header=False, index=False)
# data clean and concat end

# copy cleaned data into template

# importing openpyxl module


# opening the source excel file
source_filename = account_source_file_name_full
wb1 = xl.load_workbook(filename=source_filename)
ws1 = wb1.worksheets[0]
print(source_filename)
print(wb1.sheetnames)


# # opening the destination excel file
account_target_file_name = EGM_template + week + '.xlsx'
account_target_file_name_loc = os.path.join(EGM_teplate_loc, account_target_file_name)
print(account_target_file_name_loc)
wb2 = xl.load_workbook(filename=account_target_file_name_loc)
ws2=wb2['Machine Data - Source']
# target_list1 = os.listdir(EGM_teplate_loc)
# print(target_list1)

# calculate total number of rows and
# columns in source excel file
mr = ws1.max_row
mc = ws1.max_column

# # copying the cell values from source
# # excel file to destination excel file
for i in range(1, mr + 1):
    for j in range(1, mc + 1):
        # reading cell value from source excel file
        c = ws1.cell(row=i, column=j)

        # writing the read value to destination excel file
        ws2.cell(row=i+1, column=j).value = c.value
# saving the destination excel file
wb2.save(account_target_file_name_loc)

# workbook = pd.read_excel(account_source_file_name_full)
#
# # Creating a copy of workbook
# workbook_2 = workbook.copy()
# account_target_file_name = EGM_template + week + '.xlsx'
# account_target_file_name_loc = os.path.join(EGM_teplate_loc, account_target_file_name)
#
# with pd.ExcelWriter(account_target_file_name_loc) as writer:
#     # workbook.to_excel(writer, sheet_name='Sheet1')
#     workbook_2.to_excel(writer, sheet_name='Machine Data - Source')
